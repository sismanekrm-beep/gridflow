from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Depends, Header
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os, uuid, shutil, logging, io
import cloudinary
import cloudinary.uploader
from pathlib import Path
from pydantic import BaseModel, Field
from typing import Optional, List
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from passlib.context import CryptContext
from jose import JWTError, jwt
from google.oauth2 import id_token as google_id_token
from google.auth.transport import requests as google_requests

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

UPLOAD_DIR = ROOT_DIR / "static" / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# ─── Cloudinary Config ────────────────────────────────────────────────────────
_cld_name   = os.environ.get('CLOUDINARY_CLOUD_NAME')
_cld_key    = os.environ.get('CLOUDINARY_API_KEY')
_cld_secret = os.environ.get('CLOUDINARY_API_SECRET')
USE_CLOUDINARY = bool(_cld_name and _cld_key and _cld_secret)
if USE_CLOUDINARY:
    cloudinary.config(cloud_name=_cld_name, api_key=_cld_key, api_secret=_cld_secret, secure=True)

# ─── Auth Config ──────────────────────────────────────────────────────────────
SECRET_KEY       = os.environ.get('JWT_SECRET_KEY', 'etiket-sistemi-secret-key-2025-change-in-prod')
ALGORITHM        = "HS256"
TOKEN_EXPIRE_DAYS = 30
FREE_PRINT_LIMIT  = 3
GOOGLE_CLIENT_ID  = os.environ.get('GOOGLE_CLIENT_ID', '')

ENVELOPE_DEFAULTS = {
    "sender_address": "",
    "envelope_recipient_x": 60.0, "envelope_recipient_y": 28.0,
    "envelope_recipient_w": 172.0, "envelope_recipient_h": 65.0,
    "envelope_show_logo": True,
    "envelope_logo_x": 7.0, "envelope_logo_y": 5.0,
    "envelope_logo_w": 50.0, "envelope_logo_h": 12.0,
    "envelope_show_sender": True,
    "envelope_sender_x": 7.0, "envelope_sender_y": 18.0,
    "envelope_sender_w": 60.0, "envelope_sender_h": 18.0,
    "envelope_show_stamp": True,
    "envelope_stamp_x": 207.0, "envelope_stamp_y": 5.0,
    "envelope_stamp_w": 25.0, "envelope_stamp_h": 30.0,
}

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

def create_token(data: dict) -> str:
    payload = {**data, "exp": datetime.now(timezone.utc) + timedelta(days=TOKEN_EXPIRE_DAYS)}
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(authorization: Optional[str] = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Giriş yapmanız gerekiyor")
    try:
        payload = jwt.decode(authorization.split(" ")[1], SECRET_KEY, algorithms=[ALGORITHM])
    except JWTError:
        raise HTTPException(401, "Geçersiz veya süresi dolmuş token")
    user = await db.users.find_one({"id": payload.get("user_id")}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(401, "Kullanıcı bulunamadı")
    return user

app = FastAPI(title="Etiket Sistemi API")
api_router = APIRouter(prefix="/api")

# ─── Auth Models ─────────────────────────────────────────────────────────────
class UserRegister(BaseModel):
    email: str
    password: str
    name: str

class UserLogin(BaseModel):
    email: str
    password: str

# ─── Auth Endpoints ───────────────────────────────────────────────────────────
@api_router.post("/auth/register")
async def register(data: UserRegister):
    email = data.email.lower().strip()
    if not email or "@" not in email:
        raise HTTPException(400, "Geçerli bir e-posta giriniz")
    if len(data.password) < 6:
        raise HTTPException(400, "Şifre en az 6 karakter olmalıdır")
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(400, "Bu e-posta adresi zaten kayıtlı")
    user = {
        "id": str(uuid.uuid4()),
        "email": email,
        "name": data.name.strip(),
        "password_hash": pwd_context.hash(data.password),
        "is_premium": False,
        "print_count": 0,
        "auth_provider": "email",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(user)
    user.pop("_id", None); user.pop("password_hash", None)
    now = datetime.now(timezone.utc).isoformat()
    await db.label_formats.insert_many([
        {**p, "id": str(uuid.uuid4()), "user_id": user["id"], "created_at": now}
        for p in TANEX_PRESETS
    ] + [{**DEFAULT_ENVELOPE, "id": str(uuid.uuid4()), "user_id": user["id"], "created_at": now}])
    token = create_token({"user_id": user["id"], "is_premium": False})
    return {"access_token": token, "token_type": "bearer", "user": user}

@api_router.post("/auth/login")
async def login(data: UserLogin):
    email = data.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user:
        raise HTTPException(401, "E-posta veya şifre hatalı")
    if user.get("auth_provider") == "google" and not user.get("password_hash"):
        raise HTTPException(401, "Bu hesap Google ile oluşturulmuş. Lütfen 'Google ile Devam Et' butonunu kullanın.")
    if not pwd_context.verify(data.password, user.get("password_hash", "")):
        raise HTTPException(401, "E-posta veya şifre hatalı")
    user.pop("_id", None); user.pop("password_hash", None)
    token = create_token({"user_id": user["id"], "is_premium": user.get("is_premium", False)})
    return {"access_token": token, "token_type": "bearer", "user": user}

@api_router.get("/auth/me")
async def get_me(current_user = Depends(get_current_user)):
    return current_user

class UserProfileUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None

class UserPasswordChange(BaseModel):
    current_password: str
    new_password: str

class UserBillingUpdate(BaseModel):
    company_name: Optional[str] = None
    tax_office: Optional[str] = None
    tax_number: Optional[str] = None
    billing_address: Optional[str] = None
    billing_phone: Optional[str] = None
    billing_email: Optional[str] = None

@api_router.put("/auth/profile")
async def update_profile(data: UserProfileUpdate, current_user = Depends(get_current_user)):
    upd = {}
    if data.name and data.name.strip():
        upd["name"] = data.name.strip()
    if data.email and data.email.strip() and "@" in data.email:
        email = data.email.lower().strip()
        existing = await db.users.find_one({"email": email, "id": {"$ne": current_user["id"]}})
        if existing:
            raise HTTPException(400, "Bu e-posta adresi başka bir hesapta kullanılıyor")
        upd["email"] = email
    if not upd:
        raise HTTPException(400, "Güncellenecek alan yok")
    await db.users.update_one({"id": current_user["id"]}, {"$set": upd})
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "password_hash": 0})
    token = create_token({"user_id": user["id"], "is_premium": user.get("is_premium", False)})
    return {"user": user, "access_token": token}

@api_router.put("/auth/password")
async def change_password(data: UserPasswordChange, current_user = Depends(get_current_user)):
    user_full = await db.users.find_one({"id": current_user["id"]})
    if not pwd_context.verify(data.current_password, user_full.get("password_hash", "")):
        raise HTTPException(400, "Mevcut şifre hatalı")
    if len(data.new_password) < 6:
        raise HTTPException(400, "Yeni şifre en az 6 karakter olmalıdır")
    new_hash = pwd_context.hash(data.new_password)
    await db.users.update_one({"id": current_user["id"]}, {"$set": {"password_hash": new_hash}})
    return {"message": "Şifre başarıyla güncellendi"}

@api_router.put("/auth/billing")
async def update_billing(data: UserBillingUpdate, current_user = Depends(get_current_user)):
    upd = {k: v for k, v in data.model_dump().items() if v is not None}
    if upd:
        await db.users.update_one({"id": current_user["id"]}, {"$set": {"billing": upd}})
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "password_hash": 0})
    return user

@api_router.delete("/auth/account")
async def delete_account(current_user = Depends(get_current_user)):
    """GDPR: Delete user and ALL their data."""
    uid = current_user["id"]
    await db.products.delete_many({"user_id": uid})
    await db.categories.delete_many({"user_id": uid})
    await db.addresses.delete_many({"user_id": uid})
    await db.designs.delete_many({"user_id": uid})
    await db.label_formats.delete_many({"user_id": uid})
    await db.settings.delete_many({"user_id": uid})
    await db.users.delete_one({"id": uid})
    return {"message": "Hesap ve tüm veriler kalıcı olarak silindi"}

@api_router.post("/auth/upgrade")
async def upgrade_to_premium(current_user = Depends(get_current_user)):
    """Demo: instantly upgrade. In production, integrate with Stripe webhook."""
    await db.users.update_one({"id": current_user["id"]}, {"$set": {"is_premium": True}})
    user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "password_hash": 0})
    token = create_token({"user_id": user["id"], "is_premium": True})
    return {"access_token": token, "token_type": "bearer", "user": user}

class GoogleAuthRequest(BaseModel):
    credential: str

@api_router.post("/auth/google")
async def google_auth(data: GoogleAuthRequest):
    if not GOOGLE_CLIENT_ID:
        raise HTTPException(501, "Google OAuth yapılandırılmamış. GOOGLE_CLIENT_ID eksik.")
    try:
        idinfo = google_id_token.verify_oauth2_token(
            data.credential,
            google_requests.Request(),
            GOOGLE_CLIENT_ID
        )
    except ValueError as e:
        raise HTTPException(400, f"Geçersiz Google token: {e}")

    email = idinfo["email"].lower()
    name  = idinfo.get("name", email.split("@")[0])

    user = await db.users.find_one({"email": email})
    if not user:
        user = {
            "id": str(uuid.uuid4()),
            "email": email,
            "name": name,
            "password_hash": None,
            "is_premium": False,
            "print_count": 0,
            "auth_provider": "google",
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.users.insert_one(user)
        _now = datetime.now(timezone.utc).isoformat()
        await db.label_formats.insert_many([
            {**p, "id": str(uuid.uuid4()), "user_id": user["id"], "created_at": _now}
            for p in TANEX_PRESETS
        ] + [{**DEFAULT_ENVELOPE, "id": str(uuid.uuid4()), "user_id": user["id"], "created_at": _now}])

    user.pop("_id", None)
    user.pop("password_hash", None)
    token = create_token({"user_id": user["id"], "is_premium": user.get("is_premium", False)})
    return {"access_token": token, "token_type": "bearer", "user": user}

class PrintRecordRequest(BaseModel):
    label_count: int = 1

@api_router.post("/print/record")
async def record_print(data: PrintRecordRequest, current_user = Depends(get_current_user)):
    """Server-side print limit enforcement. Free users capped at FREE_PRINT_LIMIT."""
    if not current_user.get("is_premium", False):
        current_count = current_user.get("print_count", 0)
        if current_count >= FREE_PRINT_LIMIT:
            raise HTTPException(
                403,
                f"Ücretsiz planda yazdırma limitine ulaştınız ({FREE_PRINT_LIMIT} baskı). "
                "Premium'a geçerek sınırsız yazdırın."
            )
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$inc": {"print_count": 1}}
    )
    updated = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "password_hash": 0})
    return {
        "allowed": True,
        "print_count": updated.get("print_count", 0),
        "print_limit": None if updated.get("is_premium") else FREE_PRINT_LIMIT,
        "is_premium": updated.get("is_premium", False),
    }


DEFAULT_CATEGORY_FIELDS = [
    {"name": "Ürün Kodu",    "var": "{{code}}"},
    {"name": "Ölçü",         "var": "{{measurement}}"},
    {"name": "Açıklama",     "var": "{{description}}"},
    {"name": "Kalite",       "var": "{{quality}}"},
    {"name": "DIN Kodu",     "var": "{{standard_code}}"},
    {"name": "Adet",         "var": "{{default_qty}}"},
    {"name": "Barkod No",    "var": "{{barcode}}"},
    {"name": "Tarih",        "var": "{{print_date}}"},
]

class CategoryCreate(BaseModel):
    name: str
    image_url: Optional[str] = None
    description: Optional[str] = ""
    fields: Optional[List[dict]] = None

class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    image_url: Optional[str] = None
    description: Optional[str] = None
    fields: Optional[List[dict]] = None

class ProductCreate(BaseModel):
    code: str
    name: str
    measurement: Optional[str] = ""
    standard_code: Optional[str] = ""
    quality: Optional[str] = ""
    description: Optional[str] = ""
    default_qty: int = 1
    barcode: Optional[str] = ""
    image_url: Optional[str] = None
    category_id: Optional[str] = None
    custom_fields: Optional[dict] = None

class ProductUpdate(BaseModel):
    name: Optional[str] = None
    measurement: Optional[str] = None
    standard_code: Optional[str] = None
    quality: Optional[str] = None
    description: Optional[str] = None
    default_qty: Optional[int] = None
    barcode: Optional[str] = None
    image_url: Optional[str] = None
    category_id: Optional[str] = None
    custom_fields: Optional[dict] = None

class SettingsUpdate(BaseModel):
    brand_name: Optional[str] = None
    brand_logo_url: Optional[str] = None
    brand_logo_w: Optional[float] = None
    brand_logo_h: Optional[float] = None
    sender_address: Optional[str] = None
    # Envelope layout controls
    envelope_recipient_x: Optional[float] = None
    envelope_recipient_y: Optional[float] = None
    envelope_recipient_w: Optional[float] = None
    envelope_recipient_h: Optional[float] = None
    envelope_show_logo: Optional[bool] = None
    envelope_logo_x: Optional[float] = None
    envelope_logo_y: Optional[float] = None
    envelope_logo_w: Optional[float] = None
    envelope_logo_h: Optional[float] = None
    envelope_show_sender: Optional[bool] = None
    envelope_sender_x: Optional[float] = None
    envelope_sender_y: Optional[float] = None
    envelope_sender_w: Optional[float] = None
    envelope_sender_h: Optional[float] = None
    envelope_show_stamp: Optional[bool] = None
    envelope_stamp_x: Optional[float] = None
    envelope_stamp_y: Optional[float] = None
    envelope_stamp_w: Optional[float] = None
    envelope_stamp_h: Optional[float] = None
    margin_top: Optional[float] = None
    margin_bottom: Optional[float] = None
    margin_left: Optional[float] = None
    margin_right: Optional[float] = None
    gap_col: Optional[float] = None
    gap_row: Optional[float] = None
    print_margin_x: Optional[float] = None
    print_margin_y: Optional[float] = None
    label_font_family: Optional[str] = None
    label_default_font_size: Optional[float] = None
    envelope_font_family: Optional[str] = None
    envelope_sender_font_size: Optional[float] = None
    envelope_recipient_name_size: Optional[float] = None
    envelope_recipient_addr_size: Optional[float] = None
    envelope_size: Optional[str] = None
    envelope_has_window: Optional[bool] = None
    envelope_window_x: Optional[float] = None
    envelope_window_y: Optional[float] = None
    envelope_window_w: Optional[float] = None
    envelope_window_h: Optional[float] = None

class AddressCreate(BaseModel):
    name: str
    company: Optional[str] = ""
    address_line1: Optional[str] = ""
    address_line2: Optional[str] = ""
    district: Optional[str] = ""
    city: Optional[str] = ""
    postal_code: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    notes: Optional[str] = ""

class AddressUpdate(BaseModel):
    name: Optional[str] = None
    company: Optional[str] = None
    address_line1: Optional[str] = None
    address_line2: Optional[str] = None
    district: Optional[str] = None
    city: Optional[str] = None
    postal_code: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    notes: Optional[str] = None

class LabelFormatCreate(BaseModel):
    name: str
    label_width: float
    label_height: float
    cols: int
    rows: int
    margin_top: float = 5.5
    margin_bottom: float = 5.5
    margin_left: float = 7.0
    margin_right: float = 7.0
    gap_col: float = 2.0
    gap_row: float = 2.0
    border_radius: float = 4.0
    background: Optional[str] = "#FFFFFF"
    elements: Optional[list] = None
    design_id: Optional[str] = None
    page_width: Optional[float] = None   # Custom page width mm (for envelopes etc)
    page_height: Optional[float] = None  # Custom page height mm

class LabelFormatUpdate(BaseModel):
    name: Optional[str] = None
    label_width: Optional[float] = None
    label_height: Optional[float] = None
    cols: Optional[int] = None
    rows: Optional[int] = None
    margin_top: Optional[float] = None
    margin_bottom: Optional[float] = None
    margin_left: Optional[float] = None
    margin_right: Optional[float] = None
    gap_col: Optional[float] = None
    gap_row: Optional[float] = None
    border_radius: Optional[float] = None
    background: Optional[str] = None
    elements: Optional[list] = None
    design_id: Optional[str] = None
    page_width: Optional[float] = None
    page_height: Optional[float] = None

class DesignCreate(BaseModel):
    name: str
    elements: list
    background: Optional[str] = "#FFFFFF"

class DesignUpdate(BaseModel):
    name: Optional[str] = None
    elements: Optional[list] = None
    background: Optional[str] = None

# Tüm sayılar A4 (210×297 mm) üzerinde doğrulanmıştır:
#   cols×W + (cols-1)×gap_col + 2×margin_left  = 210
#   rows×H + (rows-1)×gap_row + 2×margin_top   = 297
TANEX_PRESETS = [
    # ref       name                          W       H      c  r  mL   mR   mT    mB    gC   gR   br
    # ── 1 sütun ──────────────────────────────────────────────────────────────────────────────────────
    {"name":"TANEX TW-2000 (210×297)",    "label_width":210.0,  "label_height":297.0,  "cols":1,"rows":1,  "margin_left":0.0,  "margin_right":0.0,  "margin_top":0.0,  "margin_bottom":0.0,  "gap_col":0.0,"gap_row":0.0, "border_radius":0.0},
    {"name":"TANEX TW-2001 (199.6×289)", "label_width":199.6,  "label_height":289.1,  "cols":1,"rows":1,  "margin_left":5.2,  "margin_right":5.2,  "margin_top":4.0,  "margin_bottom":4.0,  "gap_col":0.0,"gap_row":0.0, "border_radius":4.0},
    {"name":"TANEX TW-2002 (199.6×143)", "label_width":199.6,  "label_height":143.5,  "cols":1,"rows":2,  "margin_left":5.2,  "margin_right":5.2,  "margin_top":4.0,  "margin_bottom":4.0,  "gap_col":0.0,"gap_row":2.0, "border_radius":4.0},
    {"name":"TANEX TW-2003 (210×99)",    "label_width":210.0,  "label_height":99.0,   "cols":1,"rows":3,  "margin_left":0.0,  "margin_right":0.0,  "margin_top":0.0,  "margin_bottom":0.0,  "gap_col":0.0,"gap_row":0.0, "border_radius":4.0},
    {"name":"TANEX TW-2005 (210×148.5)", "label_width":210.0,  "label_height":148.5,  "cols":1,"rows":2,  "margin_left":0.0,  "margin_right":0.0,  "margin_top":0.0,  "margin_bottom":0.0,  "gap_col":0.0,"gap_row":0.0, "border_radius":4.0},
    {"name":"TANEX TW-2007 (210×74.25)", "label_width":210.0,  "label_height":74.25,  "cols":1,"rows":4,  "margin_left":0.0,  "margin_right":0.0,  "margin_top":0.0,  "margin_bottom":0.0,  "gap_col":0.0,"gap_row":0.0, "border_radius":4.0},
    # ── 2 sütun ──────────────────────────────────────────────────────────────────────────────────────
    {"name":"TANEX TW-2004 (99×139)",    "label_width":99.1,   "label_height":139.0,  "cols":2,"rows":2,  "margin_left":4.9,  "margin_right":4.9,  "margin_top":8.5,  "margin_bottom":8.5,  "gap_col":2.0,"gap_row":2.0, "border_radius":4.0},
    {"name":"TANEX TW-2006 (99×93)",     "label_width":99.1,   "label_height":93.1,   "cols":2,"rows":3,  "margin_left":4.9,  "margin_right":4.9,  "margin_top":6.9,  "margin_bottom":6.9,  "gap_col":2.0,"gap_row":2.0, "border_radius":4.0},
    {"name":"TANEX TW-2008 (99×67.7)",   "label_width":99.1,   "label_height":67.7,   "cols":2,"rows":4,  "margin_left":4.9,  "margin_right":4.9,  "margin_top":10.1, "margin_bottom":10.1, "gap_col":2.0,"gap_row":2.0, "border_radius":4.0},
    {"name":"TANEX TW-2009 (99×55)",     "label_width":99.1,   "label_height":55.0,   "cols":2,"rows":5,  "margin_left":4.9,  "margin_right":4.9,  "margin_top":7.0,  "margin_bottom":7.0,  "gap_col":2.0,"gap_row":2.0, "border_radius":4.0},
    {"name":"TANEX TW-2010 (99×57)",     "label_width":99.06,  "label_height":57.0,   "cols":2,"rows":5,  "margin_left":5.0,  "margin_right":5.0,  "margin_top":2.0,  "margin_bottom":2.0,  "gap_col":2.0,"gap_row":2.0, "border_radius":4.0},
    {"name":"TANEX TW-2011 (99×47)",     "label_width":99.1,   "label_height":47.0,   "cols":2,"rows":6,  "margin_left":4.9,  "margin_right":4.9,  "margin_top":2.5,  "margin_bottom":2.5,  "gap_col":2.0,"gap_row":2.0, "border_radius":4.0},
    {"name":"TANEX TW-2013 (99×42.3)",   "label_width":99.1,   "label_height":42.3,   "cols":2,"rows":7,  "margin_left":4.9,  "margin_right":4.9,  "margin_top":0.5,  "margin_bottom":0.5,  "gap_col":2.0,"gap_row":0.0, "border_radius":4.0},
    {"name":"TANEX TW-2014 (99×38.1)",   "label_width":99.1,   "label_height":38.1,   "cols":2,"rows":7,  "margin_left":4.9,  "margin_right":4.9,  "margin_top":9.2,  "margin_bottom":9.2,  "gap_col":2.0,"gap_row":2.0, "border_radius":4.0},
    {"name":"TANEX TW-2015 (99×33.9)",   "label_width":99.1,   "label_height":33.9,   "cols":2,"rows":8,  "margin_left":4.9,  "margin_right":4.9,  "margin_top":5.9,  "margin_bottom":5.9,  "gap_col":2.0,"gap_row":2.0, "border_radius":4.0},
    {"name":"TANEX TW-2016 (99×34)",     "label_width":99.1,   "label_height":34.0,   "cols":2,"rows":8,  "margin_left":4.9,  "margin_right":4.9,  "margin_top":5.5,  "margin_bottom":5.5,  "gap_col":2.0,"gap_row":2.0, "border_radius":4.0},
    # ── 3 sütun ──────────────────────────────────────────────────────────────────────────────────────
    {"name":"TANEX TW-2012 (63.5×72)",   "label_width":63.5,   "label_height":72.0,   "cols":3,"rows":4,  "margin_left":7.8,  "margin_right":7.8,  "margin_top":1.5,  "margin_bottom":1.5,  "gap_col":2.0,"gap_row":2.0, "border_radius":4.0},
    {"name":"TANEX TW-2017 (70×42.3)",   "label_width":70.0,   "label_height":42.3,   "cols":3,"rows":6,  "margin_left":0.0,  "margin_right":0.0,  "margin_top":16.6, "margin_bottom":16.6, "gap_col":0.0,"gap_row":2.0, "border_radius":4.0},
    {"name":"TANEX TW-2018 (63.5×46.6)", "label_width":63.5,   "label_height":46.6,   "cols":3,"rows":6,  "margin_left":7.8,  "margin_right":7.8,  "margin_top":3.7,  "margin_bottom":3.7,  "gap_col":2.0,"gap_row":2.0, "border_radius":4.0},
    {"name":"TANEX TW-2019 (70×37)",     "label_width":70.0,   "label_height":37.0,   "cols":3,"rows":7,  "margin_left":0.0,  "margin_right":0.0,  "margin_top":13.0, "margin_bottom":13.0, "gap_col":0.0,"gap_row":2.0, "border_radius":4.0},
    {"name":"TANEX TW-2021 (63.5×38.1)", "label_width":63.5,   "label_height":38.1,   "cols":3,"rows":7,  "margin_left":7.8,  "margin_right":7.8,  "margin_top":9.2,  "margin_bottom":9.2,  "gap_col":2.0,"gap_row":2.0, "border_radius":4.0},
    {"name":"TANEX TW-2024 (64×34)",     "label_width":64.0,   "label_height":34.0,   "cols":3,"rows":8,  "margin_left":7.0,  "margin_right":7.0,  "margin_top":5.5,  "margin_bottom":5.5,  "gap_col":2.0,"gap_row":2.0, "border_radius":4.0},
    {"name":"TANEX TW-2033 (63.5×25.4)", "label_width":63.5,   "label_height":25.4,   "cols":3,"rows":11, "margin_left":7.8,  "margin_right":7.8,  "margin_top":8.8,  "margin_bottom":8.8,  "gap_col":2.0,"gap_row":0.0, "border_radius":3.0},
    # ── 4 sütun ──────────────────────────────────────────────────────────────────────────────────────
    {"name":"TANEX TW-2032 (52.5×35)",   "label_width":52.5,   "label_height":35.0,   "cols":4,"rows":8,  "margin_left":0.0,  "margin_right":0.0,  "margin_top":8.5,  "margin_bottom":8.5,  "gap_col":0.0,"gap_row":0.0, "border_radius":4.0},
    {"name":"TANEX TW-2035 (52.5×33)",   "label_width":52.5,   "label_height":33.0,   "cols":4,"rows":9,  "margin_left":0.0,  "margin_right":0.0,  "margin_top":0.0,  "margin_bottom":0.0,  "gap_col":0.0,"gap_row":0.0, "border_radius":3.0},
    {"name":"TANEX TW-2036 (45×30)",     "label_width":45.0,   "label_height":30.0,   "cols":4,"rows":9,  "margin_left":12.0, "margin_right":12.0, "margin_top":5.5,  "margin_bottom":5.5,  "gap_col":2.0,"gap_row":2.0, "border_radius":3.0},
]

DEFAULT_ENVELOPE = {
    "name": "Zarf Etiketi (24×10.5cm)",
    "label_width": 105.0, "label_height": 240.0,
    "cols": 1, "rows": 1,
    "margin_top": 0.0, "margin_bottom": 0.0,
    "margin_left": 0.0, "margin_right": 0.0,
    "gap_col": 0.0, "gap_row": 0.0,
    "border_radius": 0.0,
    "page_width": 105.0, "page_height": 240.0,
}

# TW-2024 tek başına gereken yerler için (geriye dönük uyum)
DEFAULT_TW2024 = TANEX_PRESETS[13]  # index 13 = TW-2024

# ─── Helper ─────────────────────────────────────────────────────────
async def enrich_products_with_category(products: list):
    """Add category_image_url and category_name to products if they have a category_id."""
    cat_ids = list(set(p.get('category_id') for p in products if p.get('category_id')))
    if not cat_ids:
        return
    cats = await db.categories.find({'id': {'$in': cat_ids}}, {'_id': 0}).to_list(500)
    cat_map = {c['id']: c for c in cats}
    for p in products:
        cat_id = p.get('category_id')
        if cat_id and cat_id in cat_map:
            p['category_image_url'] = cat_map[cat_id].get('image_url')
            p['category_name'] = cat_map[cat_id].get('name')
        else:
            p.setdefault('category_image_url', None)
            p.setdefault('category_name', None)


# ─── Addresses ───────────────────────────────────────────────────────
@api_router.get("/addresses")
async def get_addresses(query: str = "", page: int = 1, limit: int = 20, current_user = Depends(get_current_user)):
    skip = (page - 1) * limit
    fq = {"user_id": current_user["id"]}
    if query.strip():
        fq = {"$or": [
            {"name":         {"$regex": query, "$options": "i"}},
            {"company":      {"$regex": query, "$options": "i"}},
            {"city":         {"$regex": query, "$options": "i"}},
            {"district":     {"$regex": query, "$options": "i"}},
            {"address_line1":{"$regex": query, "$options": "i"}},
            {"phone":        {"$regex": query, "$options": "i"}},
        ]}
    total = await db.addresses.count_documents(fq)
    items = await db.addresses.find(fq, {"_id": 0}).sort("name", 1).skip(skip).limit(limit).to_list(limit)
    return {"addresses": items, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit)}

@api_router.get("/addresses/{addr_id}")
async def get_address(addr_id: str):
    a = await db.addresses.find_one({"id": addr_id}, {"_id": 0})
    if not a: raise HTTPException(404, "Adres bulunamadı")
    return a

@api_router.post("/addresses")
async def create_address(addr: AddressCreate, current_user = Depends(get_current_user)):
    d = addr.model_dump()
    d["id"] = str(uuid.uuid4())
    d["user_id"] = current_user["id"]
    d["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.addresses.insert_one(d)
    d.pop("_id", None)
    return d

@api_router.put("/addresses/{addr_id}")
async def update_address(addr_id: str, addr: AddressUpdate):
    upd = addr.model_dump(exclude_unset=True)
    if not upd: raise HTTPException(400, "Güncellenecek alan yok")
    res = await db.addresses.update_one({"id": addr_id}, {"$set": upd})
    if res.matched_count == 0: raise HTTPException(404, "Adres bulunamadı")
    return await db.addresses.find_one({"id": addr_id}, {"_id": 0})

@api_router.delete("/addresses/{addr_id}")
async def delete_address(addr_id: str):
    res = await db.addresses.delete_one({"id": addr_id})
    if res.deleted_count == 0: raise HTTPException(404, "Adres bulunamadı")
    return {"message": "Adres silindi"}


# ─── Categories ──────────────────────────────────────────────────────
@api_router.get("/categories")
async def get_categories(current_user = Depends(get_current_user)):
    cats = await db.categories.find({"user_id": current_user["id"]}, {"_id": 0}).sort("name", 1).to_list(500)
    for cat in cats:
        cat['product_count'] = await db.products.count_documents({"category_id": cat['id']})
    return cats

@api_router.post("/categories")
async def create_category(category: CategoryCreate, current_user = Depends(get_current_user)):
    existing = await db.categories.find_one({"name": {"$regex": f"^{category.name.strip()}$", "$options": "i"}, "user_id": current_user["id"]})
    if existing:
        raise HTTPException(400, f"'{category.name}' adında kategori zaten mevcut")
    d = category.model_dump()
    d["name"] = d["name"].strip()
    d["id"] = str(uuid.uuid4())
    d["user_id"] = current_user["id"]
    d["created_at"] = datetime.now(timezone.utc).isoformat()
    if not d.get("fields"):
        d["fields"] = [
            {"id": str(uuid.uuid4()), "name": f["name"], "var": f["var"], "colorRules": []}
            for f in DEFAULT_CATEGORY_FIELDS
        ]
    await db.categories.insert_one(d)
    d.pop("_id", None)
    d['product_count'] = 0
    return d

@api_router.put("/categories/{cat_id}")
async def update_category(cat_id: str, category: CategoryUpdate):
    upd = category.model_dump(exclude_unset=True)
    if not upd:
        raise HTTPException(400, "Güncellenecek alan yok")
    res = await db.categories.update_one({"id": cat_id}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(404, "Kategori bulunamadı")
    cat = await db.categories.find_one({"id": cat_id}, {"_id": 0})
    cat['product_count'] = await db.products.count_documents({"category_id": cat_id})
    return cat

@api_router.delete("/categories/{cat_id}")
async def delete_category(cat_id: str):
    await db.products.update_many({"category_id": cat_id}, {"$unset": {"category_id": ""}})
    res = await db.categories.delete_one({"id": cat_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Kategori bulunamadı")
    return {"message": "Kategori silindi"}

@api_router.post("/categories/{cat_id}/duplicate")
async def duplicate_category(cat_id: str, current_user = Depends(get_current_user)):
    cat = await db.categories.find_one({"id": cat_id, "user_id": current_user["id"]}, {"_id": 0})
    if not cat:
        raise HTTPException(404, "Kategori bulunamadı")
    new_cat = {k: v for k, v in cat.items()}
    new_cat["id"] = str(uuid.uuid4())
    new_cat["name"] = f"{cat['name']} (Kopya)"
    new_cat["created_at"] = datetime.now(timezone.utc).isoformat()
    new_cat["fields"] = [
        {**f, "id": str(uuid.uuid4()), "colorRules": [
            {**r, "id": str(uuid.uuid4())} for r in f.get("colorRules", [])
        ]}
        for f in (cat.get("fields") or [])
    ]
    await db.categories.insert_one(new_cat)
    new_cat.pop("_id", None)
    new_cat["product_count"] = 0
    return new_cat

# ─── Products ───────────────────────────────────────────────────────
@api_router.get("/products")
async def get_products(query: str = "", page: int = 1, limit: int = 20, category_id: str = "", sort_by: str = "created_at", sort_dir: str = "desc", current_user = Depends(get_current_user)):
    skip = (page - 1) * limit
    filter_q = {"user_id": current_user["id"]}
    if query.strip():
        filter_q["$or"] = [
            {"code": {"$regex": query, "$options": "i"}},
            {"name": {"$regex": query, "$options": "i"}},
            {"measurement": {"$regex": query, "$options": "i"}},
            {"standard_code": {"$regex": query, "$options": "i"}},
            {"quality": {"$regex": query, "$options": "i"}},
        ]
    if category_id:
        filter_q["category_id"] = category_id
    total = await db.products.count_documents(filter_q)
    valid_sort = {"code", "name", "measurement", "standard_code", "quality", "default_qty", "created_at"}
    sort_field = sort_by if sort_by in valid_sort else "created_at"
    sort_direction = -1 if sort_dir.lower() == "desc" else 1
    items = await db.products.find(filter_q, {"_id": 0}).sort(sort_field, sort_direction).skip(skip).limit(limit).to_list(limit)
    await enrich_products_with_category(items)
    return {"products": items, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit)}

@api_router.get("/products/code")
async def get_product_by_code(code: str, current_user = Depends(get_current_user)):
    p = await db.products.find_one({"code": {"$regex": f"^{code}$", "$options": "i"}, "user_id": current_user["id"]}, {"_id": 0})
    if not p:
        raise HTTPException(404, f"'{code}' kodu ile ürün bulunamadı")
    await enrich_products_with_category([p])
    return p

# ─── Yardımcı: var → standart alan adı ───────────────────────────────
_VAR_TO_FIELD = {
    "{{code}}":          "code",
    "{{measurement}}":   "measurement",
    "{{description}}":   "description",
    "{{quality}}":       "quality",
    "{{default_qty}}":   "default_qty",
    "{{barcode}}":       "barcode",
    "{{standard_code}}": "standard_code",
    # {{print_date}} atlanır — otomatik
}

# ─── Excel Template Download ──────────────────────────────────────────
@api_router.get("/products/template")
async def download_template(
    category_id: Optional[str] = None,
    current_user = Depends(get_current_user)
):
    hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="0B4F8A")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    smp_fill  = PatternFill("solid", fgColor="EFF6FF")
    thin      = Side(style="thin", color="D1D5DB")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    lock_fill = PatternFill("solid", fgColor="FEF9C3")  # sarı = sabit alan

    # Kategori varsa alanlarını çek
    category = None
    cat_fields = []
    if category_id:
        category = await db.categories.find_one(
            {"id": category_id, "user_id": current_user["id"]}, {"_id": 0}
        )
        if category:
            cat_fields = [
                f for f in (category.get("fields") or [])
                if f.get("var") != "{{print_date}}"  # tarih otomatik, template'e koyma
            ]

    # Sütunları oluştur
    # Her sütun: (başlık, örnek_değer, genişlik, is_fixed, field_id_or_None)
    columns = []
    if cat_fields:
        # Kategori bazlı template
        sample_vals = {
            "{{code}}":          "TV-PH2-150",
            "{{measurement}}":   "PH2 x 150mm",
            "{{quality}}":       "Stanley",
            "{{description}}":   "Phillips yildiz tornavida",
            "{{default_qty}}":   "1",
            "{{barcode}}":       "3253560440118",
            "{{standard_code}}": "DIN965",
        }
        for f in cat_fields:
            var = f.get("var", "")
            std_key = _VAR_TO_FIELD.get(var)
            is_code = std_key == "code"
            title = f["name"] + (" *" if is_code else "")
            sample = sample_vals.get(var, "örnek değer")
            width  = 14 if is_code else 22
            columns.append((title, sample, width, is_code, f.get("id")))
        # Kategori Adı en sona (sabit, değiştirilmemeli)
        columns.append(("Kategori Adı", category["name"], 20, True, None))
    else:
        # Genel template
        columns = [
            ("Ürün Kodu *",     "TV-PH2-150",               14, True,  None),
            ("Ölçü",            "PH2 x 150mm",              16, False, None),
            ("Kalite / Marka",  "A4",                       14, False, None),
            ("DIN Kodu",        "DIN965",                   14, False, None),
            ("Açıklama",        "Phillips yildiz tornavida",26, False, None),
            ("Varsayılan Adet", "1",                        12, False, None),
            ("Barkod",          "3253560440118",            18, False, None),
            ("Kategori Adı",    "El Aletleri",              18, True,  None),
        ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ürünler"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    for ci, (hdr, smp, w, is_fixed, *_) in enumerate(columns, 1):
        col_l = get_column_letter(ci)
        h = ws.cell(row=1, column=ci, value=hdr)
        h.font = hdr_font
        h.fill = lock_fill if is_fixed else hdr_fill
        if is_fixed:
            h.font = Font(bold=True, color="92400E", size=11)
        h.alignment = hdr_align
        h.border = bdr
        s = ws.cell(row=2, column=ci, value=smp)
        s.fill = smp_fill
        s.border = bdr
        s.alignment = Alignment(vertical="center")
        ws.column_dimensions[col_l].width = w

    # Açıklama sayfası
    ws2 = wb.create_sheet("Açıklama")
    ws2.column_dimensions["A"].width = 70
    cat_name_str = category["name"] if category else "ilgili kategori adı"
    notes = [
        "ÜRÜN YÜKLEME ŞABLONU — TALİMATLAR",
        "",
        "* Ürün Kodu zorunludur. Diğer alanlar isteğe bağlıdır.",
        "  Mevcut kodla eşleşen satır GÜNCELLENIR, yoksa YENİ eklenir.",
        "",
        f"Kategori Adı sütununa tam olarak  '{cat_name_str}'  yazın.",
        "  Kategori adı yanlış veya boş ise ürün kategorisiz kalır.",
        "",
        "Sarı başlıklı sütunlar sabittir — başlık adını değiştirmeyin.",
        "Sütun sırası önemli değildir; başlık adına göre eşleştirilir.",
        "",
        "Tarih alanı otomatik doldurulur, şablona eklenmez.",
    ]
    for i, note in enumerate(notes, 1):
        c = ws2.cell(row=i, column=1, value=note)
        if i == 1:
            c.font = Font(bold=True, size=12, color="0B4F8A")

    # Metadata sayfası (import sırasında kategori alanlarını eşleştirmek için)
    if cat_fields:
        ws3 = wb.create_sheet("_meta")
        ws3.sheet_state = "hidden"
        ws3.cell(row=1, column=1, value="category_id")
        ws3.cell(row=1, column=2, value=category_id)
        ws3.cell(row=2, column=1, value="category_name")
        ws3.cell(row=2, column=2, value=category["name"])
        for ri, f in enumerate(cat_fields, 3):
            ws3.cell(row=ri, column=1, value=f.get("id",""))
            ws3.cell(row=ri, column=2, value=f.get("name",""))
            ws3.cell(row=ri, column=3, value=f.get("var",""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{category['name'].replace(' ','_')}_sablonu.xlsx" if category else "urunler_sablonu.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

# ─── Excel Import ─────────────────────────────────────────────────────
@api_router.post("/products/import")
async def import_products(
    file: UploadFile = File(...),
    current_user = Depends(get_current_user)
):
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Sadece .xlsx veya .xls dosyası desteklenir")
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, f"Excel dosyası okunamadı: {str(e)}")

    raw_headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

    # Standart alan eşleştirmeleri (başlık adı → std_key)
    STD_ALIASES = {
        "ürün kodu *": "code", "ürün kodu": "code", "urun kodu *": "code",
        "urun kodu": "code", "kod": "code",
        "ürün adı *": "name", "ürün adı": "name", "urun adi": "name",
        "ad": "name", "isim": "name",
        "ölçü": "measurement", "olcu": "measurement",
        "boyut / olcu": "measurement", "boyut/olcu": "measurement",
        "boyut": "measurement",
        "standart kodu": "standard_code", "standart": "standard_code",
        "din kodu": "standard_code", "din": "standard_code",
        "kalite": "quality", "marka": "quality", "kalite / marka": "quality",
        "kalite/marka": "quality", "marka / kalite": "quality",
        "açıklama": "description", "aciklama": "description",
        "varsayılan adet": "default_qty", "varsayilan adet": "default_qty",
        "adet": "default_qty",
        "barkod": "barcode", "barkod no": "barcode",
        "kategori adı": "category_name", "kategori adi": "category_name",
        "kategori": "category_name",
    }

    # Sabit (standart) sütun haritası: index → std_key
    std_col: dict[int, str] = {}
    # Bilinmeyen başlıklar: index → başlık (muhtemelen custom alan adı)
    unknown_cols: dict[int, str] = {}

    for i, h in enumerate(raw_headers):
        if not h:
            continue
        alias = STD_ALIASES.get(h.lower())
        if alias:
            std_col[i] = alias
        else:
            unknown_cols[i] = h  # custom alan adayı

    if "code" not in std_col.values():
        raise HTTPException(400, "Gerekli sütun bulunamadı: 'Ürün Kodu *' veya 'Ürün Kodu'")

    # Tüm kategorileri yükle (kullanıcıya ait)
    cats = await db.categories.find(
        {"user_id": current_user["id"]}, {"_id": 0}
    ).to_list(1000)
    cat_by_name: dict[str, dict] = {c["name"].lower(): c for c in cats}

    def _clean(val, default=""):
        s = str(val).strip() if val is not None else ""
        return default if s.lower() in ("nan", "none", "") else s

    def _int(val, default=1):
        try:
            v = int(float(str(val)))
            return v if v > 0 else default
        except Exception:
            return default

    results = {"added": 0, "updated": 0, "skipped": 0, "errors": []}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        def cell(i):
            return row[i] if i < len(row) else None

        # Standart alanları çıkar
        rd = {key: cell(i) for i, key in std_col.items()}

        code = _clean(rd.get("code"))
        if not code:
            results["errors"].append(f"Satır {row_idx}: Ürün kodu boş")
            results["skipped"] += 1
            continue

        name = _clean(rd.get("name")) or code  # Ürün Adı yoksa kod kullan

        # Kategoriyi bul
        cat_name_raw = _clean(rd.get("category_name"))
        category = cat_by_name.get(cat_name_raw.lower()) if cat_name_raw else None
        if cat_name_raw and not category:
            results["errors"].append(
                f"Satır {row_idx} ({code}): Kategori '{cat_name_raw}' bulunamadı — kategorisiz eklendi"
            )

        # Kategori alanlarını adıyla eşleştir
        # {alan_adı_lower → field_doc}
        cat_field_by_name: dict[str, dict] = {}
        if category:
            for f in (category.get("fields") or []):
                cat_field_by_name[f["name"].lower()] = f

        # Bilinmeyen sütunları custom alanlara eşleştir
        custom_fields: dict[str, str] = {}
        for col_i, col_header in unknown_cols.items():
            val = _clean(cell(col_i))
            if not val:
                continue
            matched_field = cat_field_by_name.get(col_header.lower())
            if matched_field:
                var = matched_field.get("var", "")
                std_key = _VAR_TO_FIELD.get(var)
                if std_key:
                    # Aslında standart bir alan (örn. template başlığı farklı yazılmış)
                    rd[std_key] = rd.get(std_key) or val
                elif var != "{{print_date}}":
                    custom_fields[matched_field["id"]] = val
            # Eşleşme yoksa sütunu sessizce atla

        pdata: dict = {
            "name":          name,
            "measurement":   _clean(rd.get("measurement")),
            "standard_code": _clean(rd.get("standard_code")),
            "quality":       _clean(rd.get("quality")),
            "description":   _clean(rd.get("description")),
            "default_qty":   _int(rd.get("default_qty")),
            "barcode":       _clean(rd.get("barcode")),
        }
        if category:
            pdata["category_id"] = category["id"]
        # standard_code değerini, kategoride aynı isimli custom alana da yaz
        # ({{standard_code}} var'ı olmayan eski kategoriler için)
        std_code_val = pdata.get("standard_code")
        if std_code_val and category:
            _din_names = {"din kodu", "standart kodu", "standart", "din"}
            for _f in (category.get("fields") or []):
                _f_var = _f.get("var", "")
                if _f["name"].lower() in _din_names and _VAR_TO_FIELD.get(_f_var) != "standard_code":
                    custom_fields[_f["id"]] = std_code_val
        if custom_fields:
            pdata["custom_fields"] = custom_fields

        existing = await db.products.find_one(
            {"code": {"$regex": f"^{code}$", "$options": "i"}, "user_id": current_user["id"]}
        )
        if existing:
            await db.products.update_one({"id": existing["id"]}, {"$set": pdata})
            results["updated"] += 1
        else:
            await db.products.insert_one({
                "id": str(uuid.uuid4()),
                "code": code,
                "image_url": None,
                "category_id": category["id"] if category else None,
                "user_id": current_user["id"],
                "created_at": datetime.now(timezone.utc).isoformat(),
                **pdata,
            })
            results["added"] += 1

    return results

@api_router.get("/products/{product_id}")
async def get_product(product_id: str):
    p = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Ürün bulunamadı")
    await enrich_products_with_category([p])
    return p

@api_router.post("/products")
async def create_product(product: ProductCreate, current_user = Depends(get_current_user)):
    existing = await db.products.find_one({"code": {"$regex": f"^{product.code.strip()}$", "$options": "i"}, "user_id": current_user["id"]})
    if existing:
        raise HTTPException(400, f"'{product.code}' kodu zaten mevcut")
    d = product.model_dump()
    d["code"] = d["code"].strip()
    d["id"] = str(uuid.uuid4())
    d["user_id"] = current_user["id"]
    d["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.products.insert_one(d)
    d.pop("_id", None)
    await enrich_products_with_category([d])
    return d

@api_router.put("/products/{product_id}")
async def update_product(product_id: str, product: ProductUpdate):
    update_data = product.model_dump(exclude_unset=True)
    if not update_data:
        raise HTTPException(400, "Güncellenecek alan yok")
    res = await db.products.update_one({"id": product_id}, {"$set": update_data})
    if res.matched_count == 0:
        raise HTTPException(404, "Ürün bulunamadı")
    updated = await db.products.find_one({"id": product_id}, {"_id": 0})
    await enrich_products_with_category([updated])
    return updated

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str):
    res = await db.products.delete_one({"id": product_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Ürün bulunamadı")
    return {"message": "Ürün başarıyla silindi"}




# ─── Design Catalog (bağımsız tasarım kütüphanesi) ──────────────────────
@api_router.get("/designs")
async def get_designs():
    designs = await db.designs.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    # Enrich each design with which formats are using it
    all_formats = await db.label_formats.find({}, {"_id": 0, "id": 1, "name": 1, "design_id": 1}).to_list(100)
    for d in designs:
        d["used_by"] = [{"id": f["id"], "name": f["name"]} for f in all_formats if f.get("design_id") == d["id"]]
    return designs

@api_router.get("/designs/{design_id}")
async def get_design(design_id: str):
    d = await db.designs.find_one({"id": design_id}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Tasarım bulunamadı")
    return d

@api_router.post("/designs")
async def create_design(design: DesignCreate):
    d = design.model_dump()
    d["id"] = str(uuid.uuid4())
    d["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.designs.insert_one(d)
    d.pop("_id", None)
    d["used_by"] = []
    return d

@api_router.put("/designs/{design_id}")
async def update_design(design_id: str, design: DesignUpdate):
    upd = design.model_dump(exclude_unset=True)
    if not upd:
        raise HTTPException(400, "Güncellenecek alan yok")
    res = await db.designs.update_one({"id": design_id}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(404, "Tasarım bulunamadı")
    d = await db.designs.find_one({"id": design_id}, {"_id": 0})
    all_formats = await db.label_formats.find({}, {"_id": 0, "id": 1, "name": 1, "design_id": 1}).to_list(100)
    d["used_by"] = [{"id": f["id"], "name": f["name"]} for f in all_formats if f.get("design_id") == design_id]
    return d

@api_router.delete("/designs/{design_id}")
async def delete_design(design_id: str):
    # Remove reference from any formats using this design
    await db.label_formats.update_many({"design_id": design_id}, {"$unset": {"design_id": ""}})
    res = await db.designs.delete_one({"id": design_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Tasarım bulunamadı")
    return {"message": "Tasarım silindi"}

@api_router.post("/designs/{design_id}/assign")
async def assign_design_to_formats(design_id: str, data: dict):
    """Assign a catalog design to one or more label formats."""
    d = await db.designs.find_one({"id": design_id})
    if not d:
        raise HTTPException(404, "Tasarım bulunamadı")
    format_ids = data.get("format_ids", [])
    # Assign to selected formats
    if format_ids:
        await db.label_formats.update_many({"id": {"$in": format_ids}}, {"$set": {"design_id": design_id}})
    return {"assigned": len(format_ids)}


# ─── Label Formats ───────────────────────────────────────────────────
@api_router.get("/label-formats")
async def get_label_formats(current_user = Depends(get_current_user)):
    fmts = await db.label_formats.find({"user_id": current_user["id"]}, {"_id": 0}).sort("created_at", 1).to_list(100)
    existing_names = {f["name"] for f in fmts}
    now = datetime.now(timezone.utc).isoformat()
    to_add = []
    for p in TANEX_PRESETS:
        if p["name"] not in existing_names:
            to_add.append({**p, "id": str(uuid.uuid4()), "user_id": current_user["id"], "created_at": now})
    if DEFAULT_ENVELOPE["name"] not in existing_names:
        to_add.append({**DEFAULT_ENVELOPE, "id": str(uuid.uuid4()), "user_id": current_user["id"], "created_at": now})
    if to_add:
        await db.label_formats.insert_many(to_add)
        for d in to_add: d.pop("_id", None)
        fmts.extend(to_add)
    return fmts

@api_router.post("/label-formats/load-presets")
async def load_tanex_presets(current_user = Depends(get_current_user)):
    """Eksik Tanex standart ölçülerini kullanıcının hesabına ekler."""
    existing = await db.label_formats.find(
        {"user_id": current_user["id"]}, {"_id": 0, "name": 1}
    ).to_list(200)
    existing_names = {f["name"] for f in existing}
    now = datetime.now(timezone.utc).isoformat()
    to_add = []
    for p in TANEX_PRESETS:
        if p["name"] not in existing_names:
            to_add.append({**p, "id": str(uuid.uuid4()), "user_id": current_user["id"], "created_at": now})
    if DEFAULT_ENVELOPE["name"] not in existing_names:
        to_add.append({**DEFAULT_ENVELOPE, "id": str(uuid.uuid4()), "user_id": current_user["id"], "created_at": now})
    if to_add:
        await db.label_formats.insert_many(to_add)
        for d in to_add: d.pop("_id", None)
    return {"added": len(to_add), "already_had": len(existing_names)}

@api_router.post("/label-formats")
async def create_label_format(fmt: LabelFormatCreate, current_user = Depends(get_current_user)):
    existing = await db.label_formats.find_one({"name": {"$regex": f"^{fmt.name.strip()}$", "$options": "i"}, "user_id": current_user["id"]})
    if existing:
        raise HTTPException(400, f"'{fmt.name}' adında format zaten mevcut")
    d = fmt.model_dump()
    d["name"] = d["name"].strip()
    d["id"] = str(uuid.uuid4())
    d["user_id"] = current_user["id"]
    d["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.label_formats.insert_one(d)
    d.pop("_id", None)
    return d

@api_router.put("/label-formats/{fmt_id}")
async def update_label_format(fmt_id: str, fmt: LabelFormatUpdate):
    upd = fmt.model_dump(exclude_unset=True)
    if not upd:
        raise HTTPException(400, "Güncellenecek alan yok")
    res = await db.label_formats.update_one({"id": fmt_id}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(404, "Format bulunamadı")
    return await db.label_formats.find_one({"id": fmt_id}, {"_id": 0})

@api_router.delete("/label-formats/{fmt_id}")
async def delete_label_format(fmt_id: str):
    count = await db.label_formats.count_documents({})
    if count <= 1:
        raise HTTPException(400, "En az bir etiket formatı olmalıdır, silemezsiniz")
    res = await db.label_formats.delete_one({"id": fmt_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Format bulunamadı")
    return {"message": "Format silindi"}


# ─── Upload ─────────────────────────────────────────────────────────
@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    if file.content_type not in ["image/jpeg", "image/png", "image/gif", "image/webp"]:
        raise HTTPException(400, "Sadece resim dosyaları desteklenir (JPG, PNG, GIF, WebP)")
    contents = await file.read()
    if USE_CLOUDINARY:
        result = cloudinary.uploader.upload(
            contents,
            folder="gridlabel",
            resource_type="image",
        )
        return {"url": result["secure_url"]}
    # Fallback: local storage
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "jpg"
    fname = f"{uuid.uuid4()}.{ext}"
    with open(UPLOAD_DIR / fname, "wb") as buf:
        buf.write(contents)
    return {"url": f"/api/static/uploads/{fname}"}

# ─── Settings ───────────────────────────────────────────────────────
@api_router.get("/settings")
async def get_settings(current_user = Depends(get_current_user)):
    s = await db.settings.find_one({"user_id": current_user["id"]}, {"_id": 0})
    if not s:
        return {
            "brand_name": "Marka Adı", "brand_logo_url": None,
            "margin_top": 5.5, "margin_bottom": 5.5,
            "margin_left": 7.0, "margin_right": 7.0,
            "gap_col": 2.0, "gap_row": 2.0,
            **ENVELOPE_DEFAULTS,
        }
    # Always merge with defaults so missing fields get default values
    return {**ENVELOPE_DEFAULTS, **s}

@api_router.put("/settings")
async def update_settings(settings: SettingsUpdate, current_user = Depends(get_current_user)):
    upd = settings.model_dump(exclude_unset=True)
    if await db.settings.find_one({"user_id": current_user["id"]}):
        await db.settings.update_one({"user_id": current_user["id"]}, {"$set": upd})
    else:
        defaults = {
            "user_id": current_user["id"],
            "brand_name": "Marka Adı", "brand_logo_url": None,
            "margin_top": 5.5, "margin_bottom": 5.5,
            "margin_left": 7.0, "margin_right": 7.0,
            "gap_col": 2.0, "gap_row": 2.0,
        }
        defaults.update(upd)
        await db.settings.insert_one(defaults)
    s = await db.settings.find_one({"user_id": current_user["id"]}, {"_id": 0})
    return {**ENVELOPE_DEFAULTS, **s}

# ─── Stats ──────────────────────────────────────────────────────────
@api_router.get("/stats")
async def get_stats(current_user = Depends(get_current_user)):
    uid = current_user["id"]
    total_products = await db.products.count_documents({"user_id": uid})
    total_categories = await db.categories.count_documents({"user_id": uid})
    recent = await db.products.find({"user_id": uid}, {"_id": 0}).sort("created_at", -1).limit(8).to_list(8)
    await enrich_products_with_category(recent)
    return {"total_products": total_products, "total_categories": total_categories, "recent_products": recent}

@api_router.get("/")
async def root():
    return {"message": "Etiket Sistemi API v2"}

# ─── App Setup ──────────────────────────────────────────────────────
app.include_router(api_router)
app.mount("/api/static", StaticFiles(directory=str(ROOT_DIR / "static")), name="static")

_cors_origins = [o.strip() for o in os.environ.get('CORS_ORIGINS', '*').split(',') if o.strip()]
_allow_credentials = '*' not in _cors_origins
app.add_middleware(
    CORSMiddleware,
    allow_credentials=_allow_credentials,
    allow_origins=_cors_origins if _allow_credentials else ["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
